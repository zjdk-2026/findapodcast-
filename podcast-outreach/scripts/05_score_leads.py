#!/usr/bin/env python3
"""
Score podcasts for outreach priority.
Usage: python scripts/05_score_leads.py [--dry-run]
"""
import sys
import argparse
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from dotenv import load_dotenv
load_dotenv(dotenv_path=Path(__file__).parent.parent / "config" / ".env")

import pandas as pd
from sqlalchemy import text
from rich.console import Console
from rich.table import Table

from src.db.connect import get_engine
from src.scoring.lead_scorer import score_all

console = Console()

TIER_COLORS = {"hot": "red", "warm": "yellow", "cold": "blue", "skip": "dim"}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Print results without saving")
    args = parser.parse_args()

    console.print("[bold cyan]Podcast Outreach — Lead Scoring[/bold cyan]\n")

    engine = get_engine()
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT id, title, host_name, category, niche_tags, total_episodes,
                   last_episode_date, contact_email, linkedin_url, instagram_url,
                   website, listen_score, country
            FROM podcasts
        """))
        rows = [dict(r) for r in result.mappings().all()]

    console.print(f"Scoring {len(rows)} podcasts...")
    scored = score_all(rows)

    tier_counts = {}
    for p in scored:
        tier = p["lead_tier"]
        tier_counts[tier] = tier_counts.get(tier, 0) + 1

    summary = Table("Tier", "Count", title="Lead Scoring Results")
    for tier in ("hot", "warm", "cold", "skip"):
        count = tier_counts.get(tier, 0)
        color = TIER_COLORS[tier]
        summary.add_row(f"[{color}]{tier.upper()}[/{color}]", str(count))
    console.print(summary)

    top_hot = [p for p in scored if p["lead_tier"] == "hot"][:10]
    if top_hot:
        console.print("\n[bold red]Top Hot Leads:[/bold red]")
        hot_table = Table("Title", "Score", "Email", "Episodes", "Category")
        for p in top_hot:
            hot_table.add_row(
                (p.get("title") or "")[:45],
                str(p["lead_score"]),
                "YES" if p.get("contact_email") else "no",
                str(p.get("total_episodes") or ""),
                (p.get("category") or "")[:20],
            )
        console.print(hot_table)

    if not args.dry_run:
        df = pd.DataFrame(scored)
        date_str = datetime.now().strftime("%Y-%m-%d")
        output_dir = Path(__file__).parent.parent / "data" / "exports"
        output_dir.mkdir(parents=True, exist_ok=True)
        out_path = output_dir / f"scored_leads_{date_str}.xlsx"

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="All Leads")

            wb = writer.book
            ws = wb["All Leads"]

            from openpyxl.styles import PatternFill
            tier_fills = {
                "hot": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
                "warm": PatternFill(start_color="FFAA00", end_color="FFAA00", fill_type="solid"),
                "cold": PatternFill(start_color="4444FF", end_color="4444FF", fill_type="solid"),
                "skip": PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"),
            }

            tier_col = None
            for i, cell in enumerate(ws[1], 1):
                if cell.value == "lead_tier":
                    tier_col = i
                    break

            if tier_col:
                for row in ws.iter_rows(min_row=2):
                    tier_cell = row[tier_col - 1]
                    fill = tier_fills.get(tier_cell.value)
                    if fill:
                        for cell in row:
                            cell.fill = fill

        console.print(f"\n[bold green]Saved:[/bold green] {out_path}")
        console.print("Next step: run [bold]python scripts/06_generate_emails.py[/bold]")
    else:
        console.print("\n[yellow]Dry run — no file saved[/yellow]")


if __name__ == "__main__":
    main()
